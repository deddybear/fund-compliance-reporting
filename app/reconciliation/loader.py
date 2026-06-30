
from __future__ import annotations
from decimal import Decimal
from pathlib import Path
from openpyxl import load_workbook
from app.reconciliation.models import ExpectedFigure
from app.reconciliation.mapping import FIGURE_RESULT_MAPPING

class ReconciliationLoader:
    """
    Load expected figures from answer key workbook.
    """

    def load(
        self,
        path: Path,
    ) -> list[ExpectedFigure]:

        if not path.exists():
            raise FileNotFoundError(
                f"Answer key not found: {path}"
            )

        workbook = load_workbook(
            filename=path,
            data_only=True,
        )

        worksheet = workbook.active

        if worksheet is None:
            workbook.close()
            raise ValueError(
                "Workbook does not contain an active worksheet."
            )

        figures: list[ExpectedFigure] = []

        for row in worksheet.iter_rows(
            min_row=2,
            values_only=True,
        ):

            if row[0] is None:
                continue

            #for mapping between excel and figure result
            metric_mapping = FIGURE_RESULT_MAPPING[str(row[1]).strip()]

            print(f"mapping metric : {metric_mapping} | {str(row[1]).strip()}")

            figures.append(
                ExpectedFigure(
                    metric_mapping=metric_mapping,
                    section=str(row[0]).strip(),
                    metric=str(row[1]).strip(),
                    value=self._parse_percentage(row[2]),
                    limit=str(row[3]).strip(),
                    utilization=str(row[4]).strip(),
                    status=str(row[5]).strip(),
                )
            )

        workbook.close()

        return figures

    @staticmethod
    def _parse_percentage(value: object) -> Decimal:
        """
        Convert values like:
            35.0%
            3.88 yrs
            SGD 38,790 / bp
        into Decimal.
        """

        text = str(value)

        text = text.replace("%", "")
        text = text.replace("yrs", "")
        text = text.replace("SGD", "")
        text = text.replace("/ bp", "")
        text = text.replace(",", "")
        text = text.strip()

        return Decimal(text)
